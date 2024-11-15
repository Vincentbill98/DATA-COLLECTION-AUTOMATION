import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import openpyxl
import csv
from datetime import datetime
import os

class RulingForm(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        tk.Label(self, text="Ruling Form").pack(pady=10)

        # Define permanent file paths
        self.file_path = "your_file_path_Ruling_Report.xlsx"
        self.csv_path = self.file_path.rsplit('.', 1)[0] + '.csv'
        self.mapping_path = "your_file_path_Ruling_Mappings.csv"

        self.load_mappings()

        self.date_var = tk.StringVar()
        self.ruling_machine_var = tk.StringVar()
        self.ruling_type_var = tk.StringVar()
        self.material_var = tk.StringVar()
        self.gsm_var = tk.StringVar()
        self.reel_kgs_var = tk.StringVar()
        self.start_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        self.end_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        self.output_reams_var = tk.StringVar()
        self.ream_packing_var = tk.StringVar()
        self.reel_size_var = tk.StringVar()
        self.cut_size_var = tk.StringVar()
        self.operator_var = tk.StringVar()
        self.remarks_var = tk.StringVar()

        frame = ttk.Frame(self)
        frame.pack(fill='both', expand=True)

        widget_frame = ttk.LabelFrame(frame, text='Insert Row')
        widget_frame.grid(row=0, column=0, padx=40, pady=30, sticky='nsew')

        ttk.Label(widget_frame, text="Date (yyyy-mm-dd):").grid(row=0, column=0, sticky='w')
        DateEntry(widget_frame, textvariable=self.date_var, date_pattern='yyyy-mm-dd').grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Ruling Machine:").grid(row=1, column=0, sticky='w')
        self.ruling_machine_cb = ttk.Combobox(widget_frame, textvariable=self.ruling_machine_var, values=self.ruling_machine_mapping)
        self.ruling_machine_cb.grid(row=1, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Ruling Type:").grid(row=2, column=0, sticky='w')
        self.ruling_type_cb = ttk.Combobox(widget_frame, textvariable=self.ruling_type_var, values=self.ruling_type_mapping)
        self.ruling_type_cb.grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Material:").grid(row=3, column=0, sticky='w')
        self.material_cb = ttk.Combobox(widget_frame, textvariable=self.material_var, values=self.material_mapping)
        self.material_cb.grid(row=3, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="GSM:").grid(row=4, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.gsm_var).grid(row=4, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Reel Kgs:").grid(row=5, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.reel_kgs_var).grid(row=5, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Start Time (HH:MM:SS):").grid(row=6, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.start_time_var).grid(row=6, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Time (HH:MM:SS):").grid(row=7, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.end_time_var).grid(row=7, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Output Reams:").grid(row=8, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.output_reams_var).grid(row=8, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Ream Packing:").grid(row=9, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.ream_packing_var).grid(row=9, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Reel Size:").grid(row=10, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.reel_size_var).grid(row=10, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Cut Size:").grid(row=11, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.cut_size_var).grid(row=11, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Operator:").grid(row=12, column=0, sticky='w')
        self.operator_cb = ttk.Combobox(widget_frame, textvariable=self.operator_var, values=self.operator_mapping)
        self.operator_cb.grid(row=12, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Remarks:").grid(row=13, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.remarks_var).grid(row=13, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(widget_frame, text="Save", command=self.insert_row).grid(row=14, column=0, columnspan=1, pady=10)
        ttk.Button(widget_frame, text="Delete Row", command=self.delete_row).grid(row=14, column=1, columnspan=1, pady=10)
        ttk.Button(widget_frame, text="Load Data", command=self.load_data_to_treeview).grid(row=15, column=0, columnspan=1, pady=10)
        ttk.Button(widget_frame, text="Toggle Theme", command=self.toggle_theme).grid(row=15, column=1, columnspan=1, pady=10)

        self.treeview = ttk.Treeview(frame, show='headings')
        self.treeview.grid(row=0, column=1, rowspan=17, sticky='nsew')

        vscrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.treeview.yview)
        hscrollbar = ttk.Scrollbar(frame, orient='horizontal', command=self.treeview.xview)
        self.treeview.configure(yscrollcommand=vscrollbar.set, xscrollcommand=hscrollbar.set)
        vscrollbar.grid(row=1, column=2, sticky='ns')
        hscrollbar.grid(row=2, column=1, sticky='ew')

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(1, weight=1)


    def load_mappings(self):
        self.ruling_machine_mapping = []
        self.ruling_type_mapping = []
        self.material_mapping = []
        self.operator_mapping = []

        if os.path.exists(self.mapping_path):
            with open(self.mapping_path, 'r') as file:
                reader = csv.reader(file)
                for row in reader:
                    if row[0] not in self.ruling_machine_mapping:
                        self.ruling_machine_mapping.append(row[0])
                    if row[1] not in self.ruling_type_mapping:
                        self.ruling_type_mapping.append(row[1])
                    if row[2] not in self.material_mapping:
                        self.material_mapping.append(row[2])
                    if row[3] not in self.operator_mapping:
                        self.operator_mapping.append(row[3])

    def save_mappings(self):
        new_mappings = [self.ruling_machine_var.get(), self.ruling_type_var.get(), self.material_var.get(), self.operator_var.get()]

        existing_mappings = []
        if os.path.exists(self.mapping_path):
            with open(self.mapping_path, 'r') as file:
                reader = csv.reader(file)
                existing_mappings = list(reader)

        if new_mappings not in existing_mappings:
            with open(self.mapping_path, 'a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(new_mappings)

    def insert_row(self):
        self.save_mappings()

        try:
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=self.date_var.get())
            sheet.cell(row=next_row, column=2, value=self.ruling_machine_var.get())
            sheet.cell(row=next_row, column=3, value=self.ruling_type_var.get())
            sheet.cell(row=next_row, column=4, value=self.material_var.get())
            sheet.cell(row=next_row, column=5, value=self.gsm_var.get())
            sheet.cell(row=next_row, column=6, value=self.reel_kgs_var.get())
            sheet.cell(row=next_row, column=7, value=self.start_time_var.get())
            sheet.cell(row=next_row, column=8, value=self.end_time_var.get())
            sheet.cell(row=next_row, column=9, value=self.output_reams_var.get())
            sheet.cell(row=next_row, column=10, value=self.ream_packing_var.get())
            sheet.cell(row=next_row, column=11, value=self.reel_size_var.get())
            sheet.cell(row=next_row, column=12, value=self.cut_size_var.get())
            sheet.cell(row=next_row, column=13, value=self.operator_var.get())
            sheet.cell(row=next_row, column=14, value=self.remarks_var.get())

            wb.save(self.file_path)

            with open(self.csv_path, 'a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow([self.date_var.get(), self.ruling_machine_var.get(), self.ruling_type_var.get(),
                                 self.material_var.get(), self.gsm_var.get(), self.reel_kgs_var.get(),
                                 self.start_time_var.get(), self.end_time_var.get(), self.output_reams_var.get(),
                                 self.ream_packing_var.get(), self.reel_size_var.get(), self.cut_size_var.get(),
                                 self.operator_var.get(), self.remarks_var.get()])

            messagebox.showinfo("Success", "Data saved successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {e}")

    def delete_row(self):
        selected_item = self.treeview.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a row to delete.")
            return

        confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this row?")
        if confirm:
            try:
                wb = openpyxl.load_workbook(self.file_path)
                sheet = wb.active

                for item in selected_item:
                    values = self.treeview.item(item, 'values')
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                        if all(str(cell.value) == str(value) for cell, value in zip(row, values)):
                            sheet.delete_rows(row[0].row, 1)
                            break

                wb.save(self.file_path)

                with open(self.csv_path, 'r') as file:
                    reader = list(csv.reader(file))

                with open(self.csv_path, 'w', newline='') as file:
                    writer = csv.writer(file)
                    for row in reader:
                        if row != list(values):
                            writer.writerow(row)
                            
                for item in selected_item:
                    self.treeview.delete(item)

                messagebox.showinfo("Success", "Row deleted successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete row: {e}")

    def load_data_to_treeview(self):
        try:
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            self.treeview["columns"] = [cell.value for cell in sheet[1]]
            for col in self.treeview["columns"]:
                self.treeview.heading(col, text=col)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.treeview.insert("", "end", values=row)

            wb.close()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")

    def toggle_theme(self):
        # Implement theme toggling logic if needed
        pass

