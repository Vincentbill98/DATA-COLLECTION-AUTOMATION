import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import openpyxl
from datetime import datetime
import tkinter.font as tkfont
import csv
import os

class FuelForm(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.file_path = "your_file_path_Vehicles.xlsx"  # Hardcoded file path

        self.vehicle_numbers = self.load_mapping("vehicle_numbers.csv")
        self.routes = self.load_mapping("routes.csv")

        insert_frame = ttk.LabelFrame(self, text='Insert Row')
        insert_frame.grid(row=0, column=0, padx=10, pady=10, sticky='nswe')

        ttk.Label(insert_frame, text="Start Date (yyyy-mm-dd):").grid(row=0, column=0, sticky='w')
        self.start_date_var = tk.StringVar()
        start_date_entry = DateEntry(insert_frame, textvariable=self.start_date_var, date_pattern='yyyy-mm-dd')
        start_date_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(insert_frame, text="Vehicle Number:").grid(row=1, column=0, sticky='w')
        self.vehicle_number_combobox = ttk.Combobox(insert_frame, values=self.vehicle_numbers)
        self.vehicle_number_combobox.set('Select Vehicle')
        self.vehicle_number_combobox.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(insert_frame, text="Mileage (Kms):").grid(row=2, column=0, sticky='w')
        self.distance_spinbox = ttk.Spinbox(insert_frame, from_=1, to=20000)
        self.distance_spinbox.insert(0, 'MILEAGE')
        self.distance_spinbox.grid(row=2, column=1, padx=5, pady=5, sticky='ew')
        self.distance_spinbox.bind('<FocusIn>', lambda e: self.distance_spinbox.delete(0, 'end'))

        ttk.Label(insert_frame, text="Fuel Quantity In:").grid(row=3, column=0, sticky='w')
        self.quantity_in_spinbox = ttk.Spinbox(insert_frame, from_=0, to=0, increment=0)
        self.quantity_in_spinbox.set('0')
        self.quantity_in_spinbox.grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        self.quantity_in_spinbox.config(state='readonly') 

        ttk.Label(insert_frame, text="Fuel Quantity Out:").grid(row=4, column=0, sticky='w')
        self.quantity_out_spinbox = ttk.Spinbox(insert_frame, from_=0, to=20000)
        self.quantity_out_spinbox.insert(0, 'FUEL QUANTITY OUT')
        self.quantity_out_spinbox.grid(row=4, column=1, padx=5, pady=5, sticky='ew')
        self.quantity_out_spinbox.bind('<FocusIn>', lambda e: self.quantity_out_spinbox.delete(0, 'end'))

        ttk.Label(insert_frame, text="Route:").grid(row=5, column=0, sticky='w')
        self.route_combobox = ttk.Combobox(insert_frame, values=self.routes)
        self.route_combobox.set('Select Route')
        self.route_combobox.grid(row=5, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(insert_frame, text="Save", command=self.insert_row).grid(row=6, column=0, pady=10, padx=10)
        ttk.Button(insert_frame, text="Load Data", command=self.load_data).grid(row=7, column=0, pady=10, padx=10)
        ttk.Button(insert_frame, text="Delete", command=self.delete_row).grid(row=8, column=0, pady=10, padx=10)

        treeview_frame = ttk.Frame(self)
        treeview_frame.grid(row=0, column=1, padx=10, pady=10, sticky='nswe')
        
        self.treeview = ttk.Treeview(treeview_frame, show='headings', selectmode='browse')
        self.treeview.pack(expand=True, fill='both')

        # Scrollbars
        vscrollbar = ttk.Scrollbar(treeview_frame, orient='vertical', command=self.treeview.yview)
        hscrollbar = ttk.Scrollbar(treeview_frame, orient='horizontal', command=self.treeview.xview)
        self.treeview.configure(yscrollcommand=vscrollbar.set, xscrollcommand=hscrollbar.set)
        vscrollbar.pack(side='right', fill='y')
        hscrollbar.pack(side='bottom', fill='x')

        # Initialize Treeview columns and headings
        self.initialize_treeview()

        # Configure column and row weights to expand the Treeview correctly
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=3)
        self.grid_rowconfigure(0, weight=1)

        self.apply_treeview_style()

        # Bind events
        self.treeview.bind('<<TreeviewSelect>>', self.on_row_select)

        # Update combo boxes with the loaded mappings
        self.update_combo_boxes()

    def initialize_treeview(self):

        self.treeview["columns"] = ["Start Date", "Vehicle Number", "Mileage (Kms)", "Fuel Quantity In", "Fuel Quantity Out", "Route"]
        self.treeview["show"] = "headings"

        for col in self.treeview["columns"]:
            self.treeview.heading(col, text=col)
            self.treeview.column(col, width=150, anchor='w')

    def apply_treeview_style(self):
        """Apply custom styles for the Treeview."""
        style = ttk.Style()
        style.configure('Treeview', rowheight=25)
        style.configure('Treeview.Heading', font=('Arial', 10, 'bold'))

        style.configure('Treeview.Selected', background='#d0eaff', foreground='black')  # Customize colors

    def on_row_select(self, event):
        """Highlight the selected row based on the current theme."""
        selected_item = self.treeview.selection()
        for item in self.treeview.get_children():
            self.treeview.item(item, tags='default')
        if selected_item:
            self.treeview.item(selected_item, tags='selected')
        self.treeview.tag_configure('default', background='white')
        self.treeview.tag_configure('selected', background='#d0eaff')  # Highlight color

    def insert_row(self):
        try:
            
            date = self.start_date_var.get()
            if not self.validate_date(date):
                raise ValueError("Invalid date format. Use yyyy-mm-dd.")
            
            vehicle_number = self.vehicle_number_combobox.get()
            if vehicle_number == 'Select Vehicle':
                raise ValueError("Please select a vehicle number.")
            
            distance = self.distance_spinbox.get()
            if not self.validate_positive_integer(distance):
                raise ValueError("Mileage must be a positive integer.")
            
            quantity_out = self.quantity_out_spinbox.get()
            if not self.validate_positive_integer(quantity_out):
                raise ValueError("Fuel quantity out must be a positive integer.")
            
            route = self.route_combobox.get()
            if route == 'Select Route':
                raise ValueError("Please select a route.")
            
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            row_values = [date, vehicle_number, distance, '0', quantity_out, route]
            sheet.append(row_values)
            workbook.save(self.file_path)

            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            with open(csv_file_path, mode='a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(row_values)

            self.treeview.insert('', 'end', values=row_values)
            self.clear_inputs()

            # Save new mappings if necessary
            updated = False
            if vehicle_number not in self.vehicle_numbers:
                self.save_mapping("vehicle_numbers.csv", [vehicle_number])
                self.vehicle_numbers.append(vehicle_number)
                updated = True
            if route not in self.routes:
                self.save_mapping("routes.csv", [route])
                self.routes.append(route)
                updated = True

            if updated:
                # Update vehicle and route combo boxes
                self.update_combo_boxes()

            messagebox.showinfo("Success", "Row inserted successfully")

        except ValueError as ve:
            messagebox.showerror("Validation Error", str(ve))
        except FileNotFoundError as fe:
            messagebox.showerror("File Error", str(fe))
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def load_data(self):
        """Load data from Excel and/or CSV into the Treeview."""
        try:
            self.treeview.delete(*self.treeview.get_children())
            
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                self.treeview.insert('', 'end', values=row)

            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            if os.path.exists(csv_file_path):
                with open(csv_file_path, mode='r') as file:
                    reader = csv.reader(file)
                    for row in reader:
                        self.treeview.insert('', 'end', values=row)

        except FileNotFoundError:
            messagebox.showerror("File Error", "The specified file does not exist.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_row(self):
        """Delete the selected row from the Treeview, Excel file, and CSV file."""
        selected_item = self.treeview.selection()
        if selected_item:
            values = self.treeview.item(selected_item, 'values')

            # Remove from Treeview
            self.treeview.delete(selected_item)

            # Remove from Excel file
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if list(row) == list(values):
                    sheet.delete_rows(row[0].row)
                    break
            workbook.save(self.file_path)

            # Remove from CSV file
            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            with open(csv_file_path, mode='r') as file:
                rows = list(csv.reader(file))
            with open(csv_file_path, mode='w', newline='') as file:
                writer = csv.writer(file)
                for row in rows:
                    if row != list(values):
                        writer.writerow(row)

            messagebox.showinfo("Success", "Row deleted successfully")

        else:
            messagebox.showwarning("Selection Error", "Please select a row to delete.")

    def validate_date(self, date_str):
        """Validate date format."""
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def validate_positive_integer(self, value):
        """Validate that the value is a positive integer."""
        try:
            return int(value) > 0
        except ValueError:
            return False

    def clear_inputs(self):
        """Clear input fields."""
        self.start_date_var.set('')
        self.vehicle_number_combobox.set('Select Vehicle')
        self.distance_spinbox.delete(0, 'end')
        self.quantity_out_spinbox.delete(0, 'end')
        self.route_combobox.set('Select Route')

    def load_mapping(self, file_name):
        """Load mapping data from a file."""
        if os.path.exists(file_name):
            with open(file_name, mode='r') as file:
                return [line.strip() for line in file]
        return []

    def save_mapping(self, file_name, new_entries):
        """Save new entries to a mapping file."""
        existing_entries = self.load_mapping(file_name)
        with open(file_name, mode='w', newline='') as file:
            writer = csv.writer(file)
            for entry in existing_entries + new_entries:
                writer.writerow([entry])

    def update_combo_boxes(self):
        """Update the values in the combo boxes with loaded mappings."""
        self.vehicle_number_combobox['values'] = self.vehicle_numbers
        self.route_combobox['values'] = self.routes

        # Provide feedback to the user
        if not self.vehicle_numbers:
            messagebox.showwarning("Warning", "No vehicle numbers found. Please ensure 'vehicle_numbers.csv' exists.")
        if not self.routes:
            messagebox.showwarning("Warning", "No routes found. Please ensure 'routes.csv' exists.")
        else:
            messagebox.showinfo("Success", "Mappings loaded successfully.")
