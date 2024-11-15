import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import json
import os
import openpyxl
import csv
import tkinter.font as tkfont  # Corrected import

class ReWindingForm(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        tk.Label(self, text="Re-Winding Form").pack(pady=10)

        self.file_path = "path/to/your/excel_file.xlsx"  # Hardcoded file path
        
        # do not change
        self.mapping_file = "path/to/your/mapping_file.json"  # Mapping file path

        self.date_var = tk.StringVar()
        self.start_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        self.end_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        self.machine_ref_var = tk.StringVar()
        self.material_type_var = tk.StringVar()
        self.reel_kg_var = tk.StringVar()
        self.reels_output_var = tk.StringVar()
        self.reel_width_var = tk.StringVar()
        self.cut_size_var = tk.StringVar()
        self.paper_gsm_var = tk.StringVar()
        self.operator_var = tk.StringVar()

        frame = ttk.Frame(self)
        frame.pack(fill='both', expand=True)

        widget_frame = ttk.LabelFrame(frame, text='Insert Row')
        widget_frame.grid(row=1, column=0, padx=40, pady=40, sticky='nsew')

        self._create_widgets(widget_frame)

        self.treeview = ttk.Treeview(frame, show='headings')
        self.treeview.grid(row=1, column=1, rowspan=14, sticky='nsew')

        vscrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.treeview.yview)
        hscrollbar = ttk.Scrollbar(frame, orient='horizontal', command=self.treeview.xview)
        self.treeview.configure(yscrollcommand=vscrollbar.set, xscrollcommand=hscrollbar.set)
        vscrollbar.grid(row=2, column=2, sticky='ns')
        hscrollbar.grid(row=3, column=1, sticky='ew')

        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(1, weight=1)

        self.load_data()
        self.populate_comboboxes()

    def _create_widgets(self, frame):
        labels = [
            ("Start Date (yyyy-mm-dd):", self.date_var),
            ("End Date (yyyy-mm-dd):", self.date_var),
            ("Start Time (HH:MM:SS):", self.start_time_var),
            ("End Time (HH:MM:SS):", self.end_time_var),
            ("Machine Reference:", self.machine_ref_var),
            ("Material Type:", self.material_type_var),
            ("Reel Kgs:", self.reel_kg_var),
            ("Reels Output:", self.reels_output_var),
            ("Reel Width:", self.reel_width_var),
            ("Cut Size:", self.cut_size_var),
            ("Paper GSM:", self.paper_gsm_var),
            ("Operator:", self.operator_var)
        ]

        for i, (text, var) in enumerate(labels):
            ttk.Label(frame, text=text).grid(row=i, column=0, sticky='w')
            entry = ttk.Entry(frame, textvariable=var)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky='ew')

        self.machine_ref_cb = ttk.Combobox(frame, textvariable=self.machine_ref_var)
        self.material_type_cb = ttk.Combobox(frame, textvariable=self.material_type_var)
        self.operator_cb = ttk.Combobox(frame, textvariable=self.operator_var)
        
        self.machine_ref_cb.grid(row=5, column=1, padx=5, pady=5, sticky='ew')
        self.material_type_cb.grid(row=6, column=1, padx=5, pady=5, sticky='ew')
        self.operator_cb.grid(row=11, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(frame, text="Save", command=self.insert_row).grid(row=len(labels), column=0, columnspan=2, pady=10)
        ttk.Button(frame, text="Delete Selected Row", command=self.delete_row).grid(row=len(labels)+1, column=0, columnspan=2, pady=10)
        ttk.Button(frame, text="Toggle Theme", command=self.toggle_theme).grid(row=len(labels)+2, column=0, columnspan=2, pady=10)

    def load_data(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file selected")
            return

        for item in self.treeview.get_children():
            self.treeview.delete(item)

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            list_values = list(sheet.values)
            if not list_values:
                return

            columns = [col for col in list_values[0] if col]

            self.treeview["columns"] = columns
            self.treeview["show"] = "headings"

            for col in columns:
                self.treeview.heading(col, text=col)
                self.treeview.column(col, anchor='w')

            for row in list_values[1:]:
                if len(row) < len(columns):
                    row = list(row) + [''] * (len(columns) - len(row))
                elif len(row) > len(columns):
                    row = row[:len(columns)]

                self.treeview.insert('', 'end', values=row)

            total_width = self.treeview.winfo_width()
            num_columns = len(columns)
            column_width = int(total_width / num_columns)

            for col_idx, col in enumerate(columns):
                max_width = tkfont.Font().measure(str(col))

                for row in list_values[1:]:
                    if col_idx < len(row):
                        cell_value = str(row[col_idx]) if col_idx < len(row) else ""
                        max_width = max(max_width, tkfont.Font().measure(cell_value))

                self.treeview.column(col, width=max(column_width, max_width + 20))

            self.treeview.update_idletasks()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def insert_row(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file selected")
            return

        try:
            start_date = self.date_var.get()
            end_date = self.date_var.get()
            start_time = self.start_time_var.get()
            end_time = self.end_time_var.get()
            machine_ref = self.machine_ref_var.get()
            material_type = self.material_type_var.get()
            reel_kg = self.reel_kg_var.get()
            reels_output = self.reels_output_var.get()
            reel_width = self.reel_width_var.get()
            cut_size = self.cut_size_var.get()
            paper_gsm = self.paper_gsm_var.get()
            operator = self.operator_var.get()

            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            row_values = [start_date, end_date, start_time, end_time, machine_ref, material_type, reel_kg, reels_output, reel_width, cut_size, paper_gsm, operator]
            sheet.append(row_values)
            workbook.save(self.file_path)

            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            with open(csv_file_path, mode='a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(row_values)

            self.treeview.insert('', 'end', values=row_values)
            self.clear_inputs()

            self.update_mappings(machine_ref, material_type, operator)

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_row(self):
        selected_items = self.treeview.selection()
        if not selected_items:
            messagebox.showerror("Error", "No row selected")
            return

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))
            rows_to_delete = []
            for item in selected_items:
                values = self.treeview.item(item, 'values')
                for i, row in enumerate(rows):
                    if row == tuple(values):
                        rows_to_delete.append(i + 1)
                        break

            rows_to_delete.sort(reverse=True)
            for row_index in rows_to_delete:
                sheet.delete_rows(row_index)

            workbook.save(self.file_path)

            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            with open(csv_file_path, mode='r', newline='') as file:
                reader = csv.reader(file)
                rows = list(reader)

            for row_index in rows_to_delete:
                rows.pop(row_index)

            with open(csv_file_path, mode='w', newline='') as file:
                writer = csv.writer(file)
                writer.writerows(rows)

            self.treeview.delete(selected_items[0])

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def populate_comboboxes(self):
        mappings = self.load_mappings()

        self.machine_ref_cb['values'] = mappings.get('machine_refs', [])
        self.material_type_cb['values'] = mappings.get('material_types', [])
        self.operator_cb['values'] = mappings.get('operators', [])

        print("Machine References:", self.machine_ref_cb['values'])
        print("Material Types:", self.material_type_cb['values'])
        print("Operators:", self.operator_cb['values'])

        if not mappings.get('machine_refs'):
            messagebox.showinfo("Info", "No machine references found in mappings.")
        if not mappings.get('material_types'):
            messagebox.showinfo("Info", "No material types found in mappings.")
        if not mappings.get('operators'):
            messagebox.showinfo("Info", "No operators found in mappings.")

    def load_mappings(self):
        if os.path.exists(self.mapping_file):
            with open(self.mapping_file, 'r') as file:
                return json.load(file)
        else:
            return {
                'machine_refs': [],
                'material_types': [],
                'operators': []
            }

    def update_mappings(self, machine_ref, material_type, operator):
        mappings = self.load_mappings()

        if machine_ref and machine_ref not in mappings['machine_refs']:
            mappings['machine_refs'].append(machine_ref)
        if material_type and material_type not in mappings['material_types']:
            mappings['material_types'].append(material_type)
        if operator and operator not in mappings['operators']:
            mappings['operators'].append(operator)

        with open(self.mapping_file, 'w') as file:
            json.dump(mappings, file, indent=4)

    def clear_inputs(self):
        self.date_var.set('')
        self.start_time_var.set(datetime.now().strftime('%H:%M:%S'))
        self.end_time_var.set(datetime.now().strftime('%H:%M:%S'))
        self.machine_ref_var.set('')
        self.material_type_var.set('')
        self.reel_kg_var.set('')
        self.reels_output_var.set('')
        self.reel_width_var.set('')
        self.cut_size_var.set('')
        self.paper_gsm_var.set('')
        self.operator_var.set('')

    def toggle_theme(self):
        # Implement your theme toggle logic here
        pass
