import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import openpyxl
from datetime import datetime
import csv

class DiecutForm(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        tk.Label(self, text="Diecut Form").pack(pady=10)

        self.file_path = "path/to/your/file.xlsx"  # Replace with your actual file path
        self.mapping_files = {
            'material_type': 'material_types.csv',
            'operator': 'operators.csv',
            'diecut_reference': 'diecut_references.csv'
        }

        self.date_var = tk.StringVar()
        self.start_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        self.end_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        self.diecut_reference_var = tk.StringVar()
        self.material_type_var = tk.StringVar()
        self.quantity_var = tk.StringVar()
        self.operator_var = tk.StringVar()

        frame = ttk.Frame(self)
        frame.pack(fill='both', expand=True)

        widget_frame = ttk.LabelFrame(frame, text='Insert row')
        widget_frame.grid(row=0, column=0, padx=40, pady=40, sticky='nsew')

        ttk.Label(widget_frame, text="Start Date (yyyy-mm-dd):").grid(row=0, column=0, sticky='w')
        self.start_date_entry = DateEntry(widget_frame, textvariable=self.date_var, date_pattern='yyyy-mm-dd')
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Date (yyyy-mm-dd):").grid(row=1, column=0, sticky='w')
        self.end_date_entry = DateEntry(widget_frame, textvariable=self.date_var, date_pattern='yyyy-mm-dd')
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Start Time (HH:MM:SS):").grid(row=2, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.start_time_var).grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Time (HH:MM:SS):").grid(row=3, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.end_time_var).grid(row=3, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Diecut Reference:").grid(row=4, column=0, sticky='w')
        self.diecut_reference_combobox = ttk.Combobox(widget_frame, textvariable=self.diecut_reference_var)
        self.diecut_reference_combobox.grid(row=4, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Material Type:").grid(row=5, column=0, sticky='w')
        self.material_type_combobox = ttk.Combobox(widget_frame, textvariable=self.material_type_var)
        self.material_type_combobox.grid(row=5, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Quantity:").grid(row=6, column=0, sticky='w')
        ttk.Entry(widget_frame, textvariable=self.quantity_var).grid(row=6, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Operator:").grid(row=7, column=0, sticky='w')
        self.operator_combobox = ttk.Combobox(widget_frame, textvariable=self.operator_var)
        self.operator_combobox.grid(row=7, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(widget_frame, text="Save", command=self.insert_row).grid(row=8, column=0, columnspan=2, pady=10)
        ttk.Button(widget_frame, text="Delete Selected Row", command=self.delete_row).grid(row=9, column=0, columnspan=2, pady=10)
        ttk.Button(widget_frame, text="Toggle Theme", command=self.toggle_theme).grid(row=10, column=0, columnspan=2, pady=10)

        self.treeview = ttk.Treeview(frame, show='headings')
        self.treeview.grid(row=0, column=1, rowspan=10, sticky='nsew')

        vscrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.treeview.yview)
        hscrollbar = ttk.Scrollbar(frame, orient='horizontal', command=self.treeview.xview)
        self.treeview.configure(yscrollcommand=vscrollbar.set, xscrollcommand=hscrollbar.set)
        vscrollbar.grid(row=2, column=2, sticky='ns')
        hscrollbar.grid(row=3, column=1, sticky='ew')

        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(1, weight=1)

        self.load_mappings()
        self.verify_mappings() 
        self.load_data()

    def load_mappings(self):
        # Load or initialize mappings for ComboBoxes
        for key, file_name in self.mapping_files.items():
            try:
                with open(file_name, 'r') as file:
                    reader = csv.reader(file)
                    data = [row[0] for row in reader if row]
                    if key == 'material_type':
                        self.material_type_combobox['values'] = data
                    elif key == 'operator':
                        self.operator_combobox['values'] = data
                    elif key == 'diecut_reference':
                        self.diecut_reference_combobox['values'] = data
            except FileNotFoundError:
                # If file doesn't exist, create an empty file
                with open(file_name, 'w') as file:
                    pass

        print("Material Types:", self.material_type_combobox['values'])
        print("Operators:", self.operator_combobox['values'])
        print("Diecut References:", self.diecut_reference_combobox['values'])

        if not self.material_type_combobox['values']:
            messagebox.showinfo("Info", "No material types found in the mappings.")
        if not self.operator_combobox['values']:
            messagebox.showinfo("Info", "No operators found in the mappings.")
        if not self.diecut_reference_combobox['values']:
            messagebox.showinfo("Info", "No diecut references found in the mappings.")

    def update_mappings(self, key, new_value):
        
        if key in self.mapping_files:
            file_name = self.mapping_files[key]
            with open(file_name, 'a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow([new_value])

    def insert_row(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file selected")
            return

        try:
            start_date = self.start_date_entry.get()
            end_date = self.end_date_entry.get()
            start_time = self.start_time_var.get()
            end_time = self.end_time_var.get()
            material_type = self.material_type_var.get()
            quantity = self.quantity_var.get()
            operator = self.operator_var.get()
            diecut_reference = self.diecut_reference_var.get()

            if material_type not in self.material_type_combobox['values']:
                self.update_mappings('material_type', material_type)
            if operator not in self.operator_combobox['values']:
                self.update_mappings('operator', operator)
            if diecut_reference not in self.diecut_reference_combobox['values']:
                self.update_mappings('diecut_reference', diecut_reference)

            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            row_values = [start_date, end_date, start_time, end_time, material_type, quantity, operator, diecut_reference]
            sheet.append(row_values)
            workbook.save(self.file_path)

            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            with open(csv_file_path, mode='a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(row_values)

            self.treeview.insert('', 'end', values=row_values)
            self.clear_inputs()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear_inputs(self):
        default_date = datetime.now()
        self.date_var.set(default_date.strftime('%Y-%m-%d'))
        default_time = datetime.now().strftime('%H:%M:%S')
        self.start_time_var.set(default_time)
        self.end_time_var.set(default_time)
        self.diecut_reference_var.set('')
        self.material_type_var.set('')
        self.quantity_var.set('')
        self.operator_var.set('')

    def delete_row(self):
        selected_items = self.treeview.selection()
        for item in selected_items:
            self.treeview.delete(item)

    def toggle_theme(self):
        current_theme = ttk.Style().theme_use()
        new_theme = 'forest-light' if current_theme == 'forest-dark' else 'forest-dark'
        ttk.Style().theme_use(new_theme)

    def load_data(self):
        self.treeview["columns"] = ["Start Date", "End Date", "Start Time", "End Time", "Material Type", "Quantity", "Operator", "Diecut Reference"]
        for col in self.treeview["columns"]:
            self.treeview.heading(col, text=col)
            self.treeview.column(col, width=100)
        self.treeview.delete(*self.treeview.get_children())

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.treeview.insert('', 'end', values=row)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def verify_mappings(self):
        if not self.material_type_combobox['values']:
            messagebox.showwarning("Warning", "Material types are missing.")
        if not self.operator_combobox['values']:
            messagebox.showwarning("Warning", "Operators are missing.")
        if not self.diecut_reference_combobox['values']:
            messagebox.showwarning("Warning", "Diecut references are missing.")
