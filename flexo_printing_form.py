import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import openpyxl
import csv
from datetime import datetime
import tkinter.font as tkfont
import os

class FlexoPrintingForm(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.file_path = "flexo_printing_data.xlsx"  # Permanent file path
        self.combo_data = {
            'operator': set(),
            'material_type': set(),
            'machine_ref': set()
        }
        self.initialize_ui()
        self.load_data()

    def initialize_ui(self):
        frame = ttk.Frame(self)
        frame.pack(fill='both', expand=True)

        widget_frame = ttk.LabelFrame(frame, text='Insert Row')
        widget_frame.grid(row=0, column=0, padx=40, pady=40, sticky='nsew')

        ttk.Label(widget_frame, text="Start Date (yyyy-mm-dd):").grid(row=0, column=0, sticky='w')
        self.start_date_entry = DateEntry(widget_frame, date_pattern='yyyy-mm-dd')
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Date (yyyy-mm-dd):").grid(row=1, column=0, sticky='w')
        self.end_date_entry = DateEntry(widget_frame, date_pattern='yyyy-mm-dd')
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Start Time (HH:MM:SS):").grid(row=2, column=0, sticky='w')
        self.start_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        ttk.Entry(widget_frame, textvariable=self.start_time_var).grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Time (HH:MM:SS):").grid(row=3, column=0, sticky='w')
        self.end_time_var = tk.StringVar(value=datetime.now().strftime('%H:%M:%S'))
        ttk.Entry(widget_frame, textvariable=self.end_time_var).grid(row=3, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(widget_frame, text="Machine Reference:").grid(row=4, column=0, sticky='w')
        self.machine_ref_var = tk.StringVar()
        self.machine_ref_cb = ttk.Combobox(widget_frame, textvariable=self.machine_ref_var, values=list(self.combo_data['machine_ref']))
        self.machine_ref_cb.grid(row=4, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Material Type:").grid(row=5, column=0, sticky='w')
        self.material_type_var = tk.StringVar()
        self.material_type_cb = ttk.Combobox(widget_frame, textvariable=self.material_type_var, values=list(self.combo_data['material_type']))
        self.material_type_cb.grid(row=5, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Reel Kgs:").grid(row=6, column=0, sticky='w')
        self.reel_kg_var = tk.StringVar()
        ttk.Entry(widget_frame, textvariable=self.reel_kg_var).grid(row=6, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Reels Output:").grid(row=7, column=0, sticky='w')
        self.reels_output_var = tk.StringVar()
        ttk.Entry(widget_frame, textvariable=self.reels_output_var).grid(row=7, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Reel Width:").grid(row=8, column=0, sticky='w')
        self.reel_width_var = tk.StringVar()
        ttk.Entry(widget_frame, textvariable=self.reel_width_var).grid(row=8, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Cut Size:").grid(row=9, column=0, sticky='w')
        self.cut_size_var = tk.StringVar()
        ttk.Entry(widget_frame, textvariable=self.cut_size_var).grid(row=9, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Paper GSM:").grid(row=10, column=0, sticky='w')
        self.paper_gsm_var = tk.StringVar()
        ttk.Entry(widget_frame, textvariable=self.paper_gsm_var).grid(row=10, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="Operator:").grid(row=11, column=0, sticky='w')
        self.operator_var = tk.StringVar()
        self.operator_cb = ttk.Combobox(widget_frame, textvariable=self.operator_var, values=list(self.combo_data['operator']))
        self.operator_cb.grid(row=11, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(widget_frame, text="Save", command=self.insert_row).grid(row=12, column=0, columnspan=2, pady=10)
        ttk.Button(widget_frame, text="Delete Selected Row", command=self.delete_row).grid(row=13, column=0, columnspan=2, pady=10)
        ttk.Button(widget_frame, text="Toggle Theme", command=self.toggle_theme).grid(row=14, column=0, columnspan=2, pady=10)

        self.treeview = ttk.Treeview(frame, show='headings')
        self.treeview.grid(row=0, column=1, rowspan=15, sticky='nsew')

        vscrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.treeview.yview)
        hscrollbar = ttk.Scrollbar(frame, orient='horizontal', command=self.treeview.xview)
        self.treeview.configure(yscrollcommand=vscrollbar.set, xscrollcommand=hscrollbar.set)
        vscrollbar.grid(row=1, column=2, sticky='ns')
        hscrollbar.grid(row=2, column=1, sticky='ew')

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(1, weight=1)

    def load_data(self):
        if not os.path.isfile(self.file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Start Time", "End Time", "Machine Reference", "Material Type", "Reel Kgs", "Reels Output", "Reel Width", "Cut Size", "Paper GSM", "Operator"])
            workbook.save(self.file_path)
        self.refresh_treeview()
        self.verify_mappings()

    def refresh_treeview(self):
        for item in self.treeview.get_children():
            self.treeview.delete(item)

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            list_values = list(sheet.values)
            if not list_values:
                return

            columns = [col for col in list_values[0] if col]

            self.treeview["columns"] = columns
            self.treeview["show"] = "headings"

            for col in columns:
                self.treeview.heading(col, text=col)
                self.treeview.column(col, anchor='w')

            for row in list_values[1:]:
                if len(row) < len(columns):
                    row = list(row) + [''] * (len(columns) - len(row))
                elif len(row) > len(columns):
                    row = row[:len(columns)]

                self.treeview.insert('', 'end', values=row)

            total_width = self.treeview.winfo_width()
            num_columns = len(columns)
            column_width = int(total_width / num_columns)

            for col_idx, col in enumerate(columns):
                max_width = tkfont.Font().measure(str(col))

                for row in list_values[1:]:
                    if col_idx < len(row):
                        cell_value = str(row[col_idx]) if col_idx < len(row) else ""
                        max_width = max(max_width, tkfont.Font().measure(cell_value))

                self.treeview.column(col, width=max(column_width, max_width + 20))

            self.treeview.update_idletasks()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def insert_row(self):
        try:
            start_date = self.start_date_entry.get()
            end_date = self.end_date_entry.get()
            start_time = self.start_time_var.get()
            end_time = self.end_time_var.get()
            machine_ref = self.machine_ref_var.get()
            material_type = self.material_type_var.get()
            reel_kg = self.reel_kg_var.get()
            reels_output = self.reels_output_var.get()
            reel_width = self.reel_width_var.get()
            cut_size = self.cut_size_var.get()
            paper_gsm = self.paper_gsm_var.get()
            operator = self.operator_var.get()

            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            sheet.append([start_date, start_time, end_time, machine_ref, material_type, reel_kg, reels_output, reel_width, cut_size, paper_gsm, operator])
            workbook.save(self.file_path)

            self.refresh_treeview()
            self.verify_mappings()
            messagebox.showinfo("Success", "Row inserted successfully")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_row(self):
        selected_item = self.treeview.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a row to delete.")
            return

        try:
            values = self.treeview.item(selected_item, 'values')
            date, start_time, end_time, machine_ref, material_type = values[:5]

            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == date and row[1] == start_time and row[2] == end_time and row[3] == machine_ref and row[4] == material_type:
                    sheet.delete_rows(row[0].row)
                    break
            workbook.save(self.file_path)
            self.refresh_treeview()
            self.verify_mappings()
            messagebox.showinfo("Success", "Row deleted successfully")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def toggle_theme(self):
        # Implement your theme toggling logic here
        pass

    def verify_mappings(self):
        print("Operator Combo Box Values:", self.operator_cb['values'])
        print("Material Type Combo Box Values:", self.material_type_cb['values'])
        print("Machine Reference Combo Box Values:", self.machine_ref_cb['values'])
