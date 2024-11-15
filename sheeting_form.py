import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import openpyxl
from datetime import datetime
import tkinter.font as tkfont
import csv
import os
import json

class SheetingForm(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.file_path = "path/to/your/file.xlsx"  # Set your permanent file path here

        tk.Label(self, text="Sheeting Form").grid(row=0, column=0, columnspan=2, pady=10)

        widget_frame = ttk.LabelFrame(self, text='Insert Row')
        widget_frame.grid(row=1, column=0, padx=10, pady=10, sticky='nswe')

        ttk.Label(widget_frame, text="Start Date (yyyy-mm-dd):").grid(row=0, column=0, sticky='w')
        self.start_date_var = tk.StringVar()
        self.start_date_entry = DateEntry(widget_frame, textvariable=self.start_date_var, date_pattern='yyyy-mm-dd')
        self.start_date_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Date (yyyy-mm-dd):").grid(row=1, column=0, sticky='w')
        self.end_date_var = tk.StringVar()
        self.end_date_entry = DateEntry(widget_frame, textvariable=self.end_date_var, date_pattern='yyyy-mm-dd')
        self.end_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky='ew')

        current_time = datetime.now().strftime('%H:%M:%S')

        ttk.Label(widget_frame, text="Start Time (HH:MM:SS):").grid(row=2, column=0, sticky='w')
        self.start_time_var = tk.StringVar(value=current_time)
        self.start_time_entry = ttk.Entry(widget_frame, textvariable=self.start_time_var)
        self.start_time_entry.grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(widget_frame, text="End Time (HH:MM:SS):").grid(row=3, column=0, sticky='w')
        self.end_time_var = tk.StringVar(value=current_time)
        self.end_time_entry = ttk.Entry(widget_frame, textvariable=self.end_time_var)
        self.end_time_entry.grid(row=3, column=1, padx=5, pady=5, sticky='ew')

        self.sheet_reference_combobox = self.create_combobox(widget_frame, 'Print Reference', 'sheet_reference')
        self.material_type_combobox = self.create_combobox(widget_frame, 'Material Type', 'material_type')
        self.operator_combobox = self.create_combobox(widget_frame, 'Operator', 'operator')

        ttk.Label(widget_frame, text="Quantity:").grid(row=6, column=0, sticky='w')
        self.quantity_spinbox = ttk.Spinbox(widget_frame, from_=0, to=10000)
        self.quantity_spinbox.insert(0, 'QUANTITY')
        self.quantity_spinbox.grid(row=6, column=1, padx=5, pady=5, sticky='ew')
        self.quantity_spinbox.bind('<FocusIn>', lambda e: self.quantity_spinbox.delete(0, 'end'))

        ttk.Button(widget_frame, text="Save", command=self.insert_row).grid(row=8, column=0, columnspan=2, pady=10)
        ttk.Button(widget_frame, text="Delete Selected Row", command=self.delete_row).grid(row=9, column=0, columnspan=2, pady=10)
        ttk.Button(widget_frame, text="Toggle Theme", command=self.toggle_theme).grid(row=10, column=0, columnspan=2, pady=10)

        self.treeview = ttk.Treeview(self, show='headings')
        self.treeview.grid(row=1, column=1, padx=10, pady=10, sticky='nswe')

        vscrollbar = ttk.Scrollbar(self, orient='vertical', command=self.treeview.yview)
        hscrollbar = ttk.Scrollbar(self, orient='horizontal', command=self.treeview.xview)
        self.treeview.configure(yscrollcommand=vscrollbar.set, xscrollcommand=hscrollbar.set)
        vscrollbar.grid(row=1, column=2, sticky='ns', rowspan=2)
        hscrollbar.grid(row=2, column=1, sticky='ew')

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.load_mappings()
        self.load_data()

    def create_combobox(self, parent, label_text, mapping_key):
        ttk.Label(parent, text=label_text + ":").grid(row=4, column=0, sticky='w')
        combobox = ttk.Combobox(parent)
        combobox.set('Select ' + label_text)
        combobox.grid(row=4, column=1, padx=5, pady=5, sticky='ew')

        combobox.bind('<FocusOut>', lambda e: self.save_mapping(mapping_key, combobox.get()))
        return combobox

    def load_data(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file path set")
            return

        for item in self.treeview.get_children():
            self.treeview.delete(item)

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            list_values = list(sheet.values)

            if not list_values:
                return 

            columns = [col for col in list_values[0] if col] 

            self.treeview["columns"] = columns
            self.treeview["show"] = "headings" 

            for col in columns:
                self.treeview.heading(col, text=col)
                self.treeview.column(col, anchor='w')

            for row in list_values[1:]:
                if len(row) < len(columns):
                    
                    row = list(row) + [''] * (len(columns) - len(row))
                elif len(row) > len(columns):
                    
                    row = row[:len(columns)]

                self.treeview.insert('', 'end', values=row)

            total_width = self.treeview.winfo_width()
            num_columns = len(columns)
            column_width = int(total_width / num_columns) 

            for col_idx, col in enumerate(columns):
                max_width = tkfont.Font().measure(str(col)) 

                for row in list_values[1:]:
                    if col_idx < len(row):
                        
                        cell_value = str(row[col_idx]) if col_idx < len(row) else ""
                        max_width = max(max_width, tkfont.Font().measure(cell_value))

                self.treeview.column(col, width=max(column_width, max_width + 20))

            self.treeview.update_idletasks() 
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def insert_row(self):
        if not self.file_path:
            messagebox.showerror("Error", "No file path set")
            return

        try:
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()
            start_time = self.start_time_var.get()
            end_time = self.end_time_var.get()
            material_type = self.material_type_combobox.get()
            quantity = self.quantity_spinbox.get()
            operator = self.operator_combobox.get()
            sheet_reference = self.sheet_reference_combobox.get()

            # Load existing data
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            # Append new row to Excel file
            row_values = [start_date, end_date, start_time, end_time, material_type, quantity, operator, sheet_reference]
            sheet.append(row_values)
            workbook.save(self.file_path)

            # Save the same data to a CSV file
            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            with open(csv_file_path, mode='a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(row_values)

            self.load_data()
            messagebox.showinfo("Success", "Row inserted successfully")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_row(self):
        selected_item = self.treeview.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Select a row to delete")
            return

        try:
            # Get selected item data
            selected_values = self.treeview.item(selected_item)['values']
            if not selected_values:
                messagebox.showwarning("Warning", "No data found for selected row")
                return

            # Remove the selected row from Treeview
            self.treeview.delete(selected_item)

            # Remove the corresponding row from Excel
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if list(row) == selected_values:
                    sheet.delete_rows(row[0].row)
                    break

            workbook.save(self.file_path)

            # Remove the corresponding row from CSV
            csv_file_path = self.file_path.rsplit('.', 1)[0] + '.csv'
            temp_file_path = csv_file_path + '.tmp'

            with open(csv_file_path, 'r') as read_file, open(temp_file_path, 'w', newline='') as write_file:
                reader = csv.reader(read_file)
                writer = csv.writer(write_file)

                for row in reader:
                    if row != list(selected_values):
                        writer.writerow(row)

            os.replace(temp_file_path, csv_file_path)

            messagebox.showinfo("Success", "Row deleted successfully")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def load_mappings(self):
        # Mapping file paths
        mappings_file = 'mappings.json'

        if not os.path.exists(mappings_file):
            messagebox.showerror("Error", "Mappings file not found")
            return

        try:
            with open(mappings_file, 'r') as file:
                mappings = json.load(file)

            # Load mappings into combo boxes
            for key in ['sheet_reference', 'material_type', 'operator']:
                values = mappings.get(key, [])
                combobox = getattr(self, f'{key}_combobox')
                combobox['values'] = values

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load mappings: {e}")

    def save_mapping(self, key, value):
        mappings_file = 'mappings.json'
        if not os.path.exists(mappings_file):
            mappings = {}
        else:
            with open(mappings_file, 'r') as file:
                mappings = json.load(file)

        if key not in mappings:
            mappings[key] = []

        if value and value not in mappings[key]:
            mappings[key].append(value)

        try:
            with open(mappings_file, 'w') as file:
                json.dump(mappings, file)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save mapping: {e}")

    def toggle_theme(self):
        # Placeholder for theme toggling
        pass
